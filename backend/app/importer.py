"""Разбор файла со списком устройств.

Файл приносят такой, какой есть: колонки называются по-разному, часть
данных отсутствует, лишние столбцы (инвентарный номер, серийник) никому
здесь не нужны, но и выбрасывать их нельзя — человек по ним опознаёт
строку. Поэтому разбор нарочно нестрогий: известные заголовки
раскладываются по полям, остальные складываются в `extra` как есть.

Ничего не проверяется и никуда не сохраняется: задача — прочитать. Годится
ли модель, существует ли группа, не занят ли IP — выясняется потом, при
переносе строки в спецификацию.
"""

import csv
import io
import re
from dataclasses import dataclass, field
from typing import Any

from app import siemens

# Синонимы заголовков. Сравнение идёт по «сжатой» форме: строчными, без
# пробелов, дефисов и точек, — чтобы «IP-адрес», «ip адрес» и «IP.адрес»
# считались одним и тем же.
HEADERS = {
    "name": ["имя", "название", "наименование", "устройство", "name", "hostname", "хост"],
    "template_name": ["модель", "шаблон", "оборудование", "template", "model"],
    "type_name": ["тип", "категория", "type", "kind"],
    "management_ip": ["ip", "ipадрес", "адрес", "managementip", "адресуправления"],
    "location": ["расположение", "место", "локация", "цех", "location", "site", "помещение"],
    "notes": ["заметки", "заметка", "примечание", "примечания", "комментарий", "notes", "comment"],
    "group_name": ["группа", "group"],
    "tags_text": ["теги", "тег", "метки", "tags"],
}

FIELDS = list(HEADERS)


@dataclass
class ParsedRow:
    """Строка файла, разложенная по известным полям."""
    row_number: int
    values: dict[str, str] = field(default_factory=dict)
    extra: dict[str, Any] = field(default_factory=dict)

    @property
    def is_empty(self) -> bool:
        return not any(self.values.values()) and not any(str(v).strip() for v in self.extra.values())


class ImportError_(Exception):
    """Файл не удалось прочитать — с текстом, который можно показать."""


def parse(filename: str, content: bytes) -> list[ParsedRow]:
    """Прочитать файл целиком. Формат определяется по расширению."""
    lowered = filename.lower()
    if lowered.endswith(".xml"):
        return _read_siemens(content)
    if lowered.endswith((".xlsx", ".xlsm")):
        table = _read_xlsx(content)
    elif lowered.endswith((".csv", ".txt")):
        table = _read_csv(content)
    else:
        raise ImportError_("Поддерживаются файлы .xlsx, .csv и .xml (выгрузка Siemens Automation Tool)")

    if not table:
        raise ImportError_("Файл пуст")

    header, *body = table
    mapping = _map_headers(header)
    rows = []
    for index, raw in enumerate(body, start=2):  # 1-я строка — заголовок
        values = {name: "" for name in FIELDS}
        extra: dict[str, Any] = {}
        for column, cell in enumerate(raw):
            text = "" if cell is None else str(cell).strip()
            target = mapping.get(column)
            if target:
                values[target] = text
            elif text:
                title = header[column] if column < len(header) else f"столбец {column + 1}"
                extra[str(title).strip() or f"столбец {column + 1}"] = text
        parsed = ParsedRow(row_number=index, values=values, extra=extra)
        if not parsed.is_empty:
            rows.append(parsed)
    return rows


def _read_siemens(content: bytes) -> list[ParsedRow]:
    """Выгрузка Siemens Automation Tool.

    У неё нет ни заголовков, ни столбцов — раскладывать по полям нечего,
    структура известна заранее. Поэтому она разбирается отдельно и приходит
    сюда уже готовыми строками.
    """
    if not siemens.looks_like_siemens(content):
        raise ImportError_("Это не похоже на XML — проверьте файл")
    try:
        found = siemens.parse_devices(content, _decode)
    except ValueError as exc:
        raise ImportError_(str(exc)) from None

    rows = []
    for index, item in enumerate(found, start=1):
        values = {name: "" for name in FIELDS}
        values.update(item["values"])
        parsed = ParsedRow(row_number=index, values=values, extra=item["extra"])
        if not parsed.is_empty:
            rows.append(parsed)
    return rows


def _map_headers(header: list[Any]) -> dict[int, str]:
    """Номер столбца -> поле. Что не опознали, остаётся в extra."""
    mapping: dict[int, str] = {}
    taken: set[str] = set()
    for column, title in enumerate(header):
        key = _squash(title)
        if not key:
            continue
        for name, synonyms in HEADERS.items():
            if name in taken:
                continue
            if key in synonyms:
                mapping[column] = name
                taken.add(name)
                break
    return mapping


def _squash(title: Any) -> str:
    return re.sub(r"[\s\-_.]+", "", str(title or "").strip().lower())


def _read_csv(content: bytes) -> list[list[Any]]:
    text = _decode(content)
    # Разделитель бывает и точкой с запятой: так Excel сохраняет CSV на
    # русской локали, и такой файл приносят чаще всего.
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    return [row for row in csv.reader(io.StringIO(text), dialect)]


def _decode(content: bytes) -> str:
    # Тот же Excel сохраняет в windows-1251, а выгрузки из систем — в utf-8.
    for encoding in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ImportError_("Не удалось определить кодировку файла — сохраните его в UTF-8")


def _read_xlsx(content: bytes) -> list[list[Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - зависимость есть в requirements
        raise ImportError_("Чтение xlsx недоступно на сервере") from exc

    try:
        book = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ImportError_("Не удалось прочитать книгу Excel — проверьте файл") from exc
    sheet = book.active
    return [list(row) for row in sheet.iter_rows(values_only=True)]
